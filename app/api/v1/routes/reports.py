from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from app.api import deps
from app.models.core import validate_path
from app.services.report_service import ReportService
from app.schemas.report import DailyCountSummary, MonthlyFinancialSummary, AttendanceTrend
from app.models.user import User
from typing import List, Optional
from datetime import date, timedelta

router = APIRouter()


def _resolve_scope(
    current_user: User,
    scope_path: Optional[str],
) -> str:
    """
    Validate and return the effective scope path.

    Replaces the original async helper that made a raw SQL round-trip
    (SELECT CAST ... <@ CAST ...) on every report request.
    deps.resolve_scope_path() performs the same check in pure Python
    using the ltree path already loaded from the JWT claim, with no
    extra database call.

    Also validates that any caller-supplied scope_path is a well-formed
    ltree string before it reaches raw SQL templates.
    """
    if scope_path and not validate_path(scope_path):
        raise HTTPException(status_code=400, detail="Invalid scope path format")
    return deps.resolve_scope_path(current_user, scope_path)

@router.get(
    "/export/csv",
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def export_report_csv(
    report_type: str = Query(..., description="counts, financial, or attendance"),
    scope_path: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Export report data as CSV.
    """
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    effective_scope = _resolve_scope(current_user, scope_path)

    if report_type == "counts":
        buffer = await ReportService.export_counts_csv(db, effective_scope, start_date, end_date)
        filename = f"counts_{start_date}_{end_date}.csv"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    if report_type == "financial":
        buffer = await ReportService.export_financial_csv(db, effective_scope, start_date, end_date)
        filename = f"financial_{start_date}_{end_date}.csv"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    if report_type == "attendance":
        buffer = await ReportService.export_attendance_csv(db, effective_scope, start_date, end_date)
        filename = f"attendance_{start_date}_{end_date}.csv"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    raise HTTPException(status_code=400, detail="Export type not supported")

@router.get(
    "/summary",
    response_model=List[DailyCountSummary],
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def get_summary_report(
    scope_path: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Get daily count summaries.
    Filtered by user's scope unless overridden by a more specific scope (if allowed).
    """
    # Default date range: last 30 days
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    effective_scope = _resolve_scope(current_user, scope_path)

    # We use user's scope as base to restrict access
    return await ReportService.get_daily_counts(db, effective_scope, start_date, end_date)

@router.get(
    "/financial",
    response_model=List[MonthlyFinancialSummary],
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def get_financial_report(
    scope_path: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Get monthly financial summary.
    """
    if not start_date:
        start_date = date.today().replace(day=1) - timedelta(days=365) # Last year
    if not end_date:
        end_date = date.today()

    effective_scope = _resolve_scope(current_user, scope_path)

    return await ReportService.get_financial_summary(db, effective_scope, start_date, end_date)

@router.get(
    "/attendance",
    response_model=List[AttendanceTrend],
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def get_attendance_report(
    scope_path: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Get worker attendance trends.
    """
    if not start_date:
        start_date = date.today() - timedelta(weeks=12) # Last 12 weeks
    if not end_date:
        end_date = date.today()

    effective_scope = _resolve_scope(current_user, scope_path)

    return await ReportService.get_attendance_trends(db, effective_scope, start_date, end_date)

@router.post("/refresh", dependencies=[Depends(deps.PermissionChecker("reports:refresh"))])
async def refresh_reports(
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Manually refresh report views (Admin only ideally).
    """
    # Permission enforced via PermissionChecker ("reports:refresh")
    await ReportService.refresh_views(db)
    return {"status": "success", "message": "Materialized views refreshed"}


# Advanced Analytics Routes
@router.get(
    "/timeseries",
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def get_timeseries_analysis(
    metric: str = Query(..., description="counts, offerings, or attendance"),
    interval: str = Query("daily", description="daily, weekly, or monthly"),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Get time series analysis for a specific metric.

    Returns data points over time with trend analysis.
    """
    if not start_date:
        start_date = date.today() - timedelta(days=90)
    if not end_date:
        end_date = date.today()

    effective_scope = str(current_user.path)

    interval_map = {
        "daily": "day",
        "weekly": "week",
        "monthly": "month",
    }
    interval_key = interval.strip().lower()
    if interval_key not in interval_map:
        raise HTTPException(status_code=400, detail="Invalid interval")

    from app.models.counts import Count
    from app.models.offerings import Offering
    from app.models.attendance import WorkerAttendance
    from app.models.programs import ProgramEvent
    from sqlalchemy import select, func, text, cast, DateTime

    def _trend_from_series(series: List[dict]) -> str:
        if len(series) < 2:
            return "stable"
        first = series[0]["value"]
        last = series[-1]["value"]
        if first == 0:
            return "up" if last > 0 else "stable"
        change = (last - first) / abs(first)
        if change > 0.05:
            return "up"
        if change < -0.05:
            return "down"
        return "stable"

    if metric == "counts":
        period = func.date_trunc(interval_map[interval_key], Count.date).label("period")
        query = select(
            period,
            func.sum(Count.adult_male + Count.adult_female + Count.youth_male +
                    Count.youth_female + Count.boys + Count.girls).label('total')
        ).where(
            Count.date.between(start_date, end_date),
            text("CAST(path AS ltree) <@ CAST(:scope_path AS ltree)").bindparams(scope_path=effective_scope)
        ).group_by(period).order_by(period)

        result = await db.execute(query)
        data = [{"date": str(row.period.date()), "value": row.total} for row in result]

        return {
            "metric": metric,
            "interval": interval_key,
            "data": data,
            "trend": _trend_from_series(data)
        }
    if metric == "offerings":
        period = func.date_trunc(interval_map[interval_key], Offering.date).label("period")
        query = select(
            period,
            func.coalesce(func.sum(Offering.amount), 0).label("total")
        ).where(
            Offering.date.between(start_date, end_date),
            text("CAST(path AS ltree) <@ CAST(:scope_path AS ltree)").bindparams(scope_path=effective_scope)
        ).group_by(period).order_by(period)
        result = await db.execute(query)
        data = [{"date": str(row.period.date()), "value": float(row.total)} for row in result]
        return {"metric": metric, "interval": interval_key, "data": data, "trend": _trend_from_series(data)}
    if metric == "attendance":
        period = func.date_trunc(
            interval_map[interval_key],
            cast(ProgramEvent.date, DateTime),
        ).label("period")
        query = select(
            period,
            func.count(WorkerAttendance.id).label("total")
        ).join(ProgramEvent, ProgramEvent.id == WorkerAttendance.event_id).where(
            ProgramEvent.date.between(start_date, end_date),
            text("CAST(path AS ltree) <@ CAST(:scope_path AS ltree)").bindparams(scope_path=effective_scope)
        ).group_by(period).order_by(period)
        result = await db.execute(query)
        data = [{"date": str(row.period.date()), "value": row.total} for row in result]
        return {"metric": metric, "interval": interval_key, "data": data, "trend": _trend_from_series(data)}

    raise HTTPException(status_code=400, detail="Metric not supported")


@router.get(
    "/by-level",
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def get_hierarchical_breakdown(
    metric: str = Query(..., description="counts, offerings, or attendance"),
    level: str = Query(..., description="location, group, region, or state"),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Get hierarchical breakdown by organizational level.

    Shows aggregated metrics for each unit at the specified level.
    """
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    effective_scope = str(current_user.path)

    from sqlalchemy import select, func, text
    from app.models.counts import Count
    from app.models.offerings import Offering
    from app.models.attendance import WorkerAttendance
    from app.models.programs import ProgramEvent

    level_map = {
        "location": 6,
        "group": 5,
        "region": 4,
        "state": 3,
    }
    level_key = level.strip().lower()
    if level_key not in level_map:
        raise HTTPException(status_code=400, detail="Invalid level")
    segment_count = level_map[level_key]

    def group_expr(model):
        return func.subpath(model.path, 0, segment_count).label("group_path")

    if metric == "counts":
        query = select(
            group_expr(Count),
            func.sum(Count.adult_male + Count.adult_female + Count.youth_male +
                    Count.youth_female + Count.boys + Count.girls).label("total")
        ).where(
            Count.date.between(start_date, end_date),
            text("CAST(path AS ltree) <@ CAST(:scope_path AS ltree)").bindparams(scope_path=effective_scope)
        ).group_by(group_expr(Count)).order_by(group_expr(Count))
        result = await db.execute(query)
        return {"metric": metric, "level": level, "breakdown": [{"path": row.group_path, "total": row.total} for row in result]}

    if metric == "offerings":
        query = select(
            group_expr(Offering),
            func.coalesce(func.sum(Offering.amount), 0).label("total")
        ).where(
            Offering.date.between(start_date, end_date),
            text("CAST(path AS ltree) <@ CAST(:scope_path AS ltree)").bindparams(scope_path=effective_scope)
        ).group_by(group_expr(Offering)).order_by(group_expr(Offering))
        result = await db.execute(query)
        return {"metric": metric, "level": level, "breakdown": [{"path": row.group_path, "total": str(row.total)} for row in result]}

    if metric == "attendance":
        query = select(
            group_expr(WorkerAttendance),
            func.count(WorkerAttendance.id).label("total")
        ).join(ProgramEvent, ProgramEvent.id == WorkerAttendance.event_id).where(
            ProgramEvent.date.between(start_date, end_date),
            text("CAST(path AS ltree) <@ CAST(:scope_path AS ltree)").bindparams(scope_path=effective_scope)
        ).group_by(group_expr(WorkerAttendance)).order_by(group_expr(WorkerAttendance))
        result = await db.execute(query)
        return {"metric": metric, "level": level, "breakdown": [{"path": row.group_path, "total": row.total} for row in result]}

    raise HTTPException(status_code=400, detail="Metric not supported")


@router.get(
    "/anomalies",
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def detect_anomalies(
    metric: str = Query("counts", description="Metric to analyze"),
    threshold: float = Query(2.0, description="Standard deviations for anomaly detection"),
    days: int = Query(30, description="Days to analyze"),
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Detect anomalies in data using statistical analysis.

    Identifies unusual patterns or outliers that may need attention.
    """
    effective_scope = str(current_user.path)
    start_date = date.today() - timedelta(days=days)

    # Simplified anomaly detection
    from app.models.counts import Count
    from sqlalchemy import select, func, text

    # Get daily totals
    query = select(
        Count.date,
        Count.location_id,
        func.sum(Count.adult_male + Count.adult_female + Count.youth_male +
                Count.youth_female + Count.boys + Count.girls).label('total')
    ).where(
        Count.date >= start_date,
        text("CAST(path AS ltree) <@ CAST(:scope_path AS ltree)").bindparams(scope_path=effective_scope)
    ).group_by(Count.date, Count.location_id)

    result = await db.execute(query)
    data = [{"date": str(row.date), "location": row.location_id, "value": row.total} for row in result]

    # Simple threshold-based detection (can be enhanced with statistical methods)
    if data:
        values = [d['value'] for d in data]
        avg = sum(values) / len(values)
        anomalies = [d for d in data if abs(d['value'] - avg) > (avg * 0.5)]  # 50% deviation

        return {
            "metric": metric,
            "period_days": days,
            "average": avg,
            "anomalies_detected": len(anomalies),
            "anomalies": anomalies[:10]  # Top 10
        }

    return {"anomalies": []}


@router.get(
    "/growth-rate",
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def get_growth_rate(
    metric: str = Query("counts", description="Metric to analyze"),
    period: str = Query("monthly", description="daily, weekly, or monthly"),
    months: int = Query(12, description="Number of months to analyze"),
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Calculate growth rate over time.

    Shows percentage change in metrics period-over-period.
    """
    effective_scope = str(current_user.path)
    start_date = date.today() - timedelta(days=months * 30)

    from app.models.counts import Count
    from sqlalchemy import select, func, text, extract

    # Monthly aggregation
    query = select(
        extract('year', Count.date).label('year'),
        extract('month', Count.date).label('month'),
        func.sum(Count.adult_male + Count.adult_female + Count.youth_male +
                Count.youth_female + Count.boys + Count.girls).label('total')
    ).where(
        Count.date >= start_date,
        text("CAST(path AS ltree) <@ CAST(:scope_path AS ltree)").bindparams(scope_path=effective_scope)
    ).group_by('year', 'month').order_by('year', 'month')

    result = await db.execute(query)
    data = [{"year": int(row.year), "month": int(row.month), "total": row.total} for row in result]

    # Calculate growth rates
    growth_rates = []
    for i in range(1, len(data)):
        prev = data[i-1]['total']
        curr = data[i]['total']
        if prev > 0:
            growth = ((curr - prev) / prev) * 100
            growth_rates.append({
                "period": f"{data[i]['year']}-{data[i]['month']:02d}",
                "value": curr,
                "growth_rate": round(growth, 2)
            })

    return {
        "metric": metric,
        "period": period,
        "data": growth_rates
    }


# Export Format Routes
@router.post(
    "/export/excel",
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def export_excel(
    report_type: str = Query(..., description="counts, financial, or attendance"),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Export report as Excel file.

    Note: Requires openpyxl package. Install with: pip install openpyxl
    """
    try:
        import openpyxl
        from io import BytesIO
    except ImportError:
        raise HTTPException(status_code=501, detail="Excel export not available. Install openpyxl.")

    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    effective_scope = str(current_user.path)

    # Get data (reuse existing service)
    if report_type == "counts":
        data = await ReportService.get_daily_counts(db, effective_scope, start_date, end_date)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Counts Report"

        # Headers
        ws.append([
            "Date",
            "Location",
            "Total Attendance",
            "Men",
            "Women",
            "Youth Male",
            "Youth Female",
            "Boys",
            "Girls",
            "Record Count",
        ])

        # Data
        for item in data:
            ws.append([
                str(item.day),
                item.location_name or "N/A",
                item.total_attendance,
                item.total_men,
                item.total_women,
                item.total_youth_male,
                item.total_youth_female,
                item.total_boys,
                item.total_girls,
                item.record_count,
            ])

        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"counts_{start_date}_{end_date}.xlsx"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    if report_type == "financial":
        data = await ReportService.get_financial_summary(db, effective_scope, start_date, end_date)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"
        ws.append(["Month", "Location", "Total Amount", "Transactions"])
        for item in data:
            ws.append([str(item.month), item.location_name or "N/A", item.total_amount, item.transaction_count])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"financial_{start_date}_{end_date}.xlsx"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    if report_type == "attendance":
        data = await ReportService.get_attendance_trends(db, effective_scope, start_date, end_date)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        ws.append(["Week", "Location", "Status", "Worker Count"])
        for item in data:
            ws.append([str(item.week), item.location_name or "N/A", item.status, item.worker_count])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"attendance_{start_date}_{end_date}.xlsx"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    raise HTTPException(status_code=400, detail="Report type not supported")

@router.post(
    "/export/pdf",
    dependencies=[Depends(deps.PermissionChecker("reports:read"))],
)
async def export_pdf(
    report_type: str = Query(..., description="counts, financial, or attendance"),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_active_user),
):
    """
    Export report as PDF file.

    Note: Requires reportlab package. Install with: pip install reportlab
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from io import BytesIO
    except ImportError:
        raise HTTPException(status_code=501, detail="PDF export not available. Install reportlab.")

    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    effective_scope = str(current_user.path)

    # Get data
    if report_type == "counts":
        data = await ReportService.get_daily_counts(db, effective_scope, start_date, end_date)

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        # Title
        styles = getSampleStyleSheet()
        title = Paragraph(f"Counts Report: {start_date} to {end_date}", styles['Title'])
        elements.append(title)

        # Table data
        table_data = [[
            "Date",
            "Location",
            "Total Attendance",
            "Men",
            "Women",
            "Youth Male",
            "Youth Female",
            "Boys",
            "Girls",
            "Record Count",
        ]]
        for item in data[:50]:  # Limit to 50 rows for PDF
            table_data.append([
                str(item.day),
                item.location_name or "N/A",
                str(item.total_attendance),
                str(item.total_men),
                str(item.total_women),
                str(item.total_youth_male),
                str(item.total_youth_female),
                str(item.total_boys),
                str(item.total_girls),
                str(item.record_count),
            ])

        # Create table
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)

        # Build PDF
        doc.build(elements)
        buffer.seek(0)

        filename = f"counts_{start_date}_{end_date}.pdf"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    if report_type == "financial":
        data = await ReportService.get_financial_summary(db, effective_scope, start_date, end_date)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        title = Paragraph(f"Financial Report: {start_date} to {end_date}", styles['Title'])
        elements.append(title)
        table_data = [["Month", "Location", "Total Amount", "Transactions"]]
        for item in data[:50]:
            table_data.append([str(item.month), item.location_name or "N/A", str(item.total_amount), str(item.transaction_count)])
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        filename = f"financial_{start_date}_{end_date}.pdf"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    if report_type == "attendance":
        data = await ReportService.get_attendance_trends(db, effective_scope, start_date, end_date)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        title = Paragraph(f"Attendance Report: {start_date} to {end_date}", styles['Title'])
        elements.append(title)
        table_data = [["Week", "Location", "Status", "Worker Count"]]
        for item in data[:50]:
            table_data.append([str(item.week), item.location_name or "N/A", item.status, str(item.worker_count)])
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        filename = f"attendance_{start_date}_{end_date}.pdf"
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    raise HTTPException(status_code=400, detail="Report type not supported")
